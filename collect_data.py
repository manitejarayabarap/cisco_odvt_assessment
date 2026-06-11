"""
collect_data.py

Sweeps a device under test through a grid of parameter settings,
collects signal characterisation measurements for each, and saves
results to both a CSV file and a formatted Excel workbook.

--- CHANGES FROM ORIGINAL ---
Strategy: Two-phase adaptive sweep.

Phase 1 - Coarse scan: Sample a sparse grid across the full parameter
space to map the performance landscape quickly. Uses fewer samples
(n_samples=100) to reduce measurement time per point.

Phase 2 - Focused fine scan: Identify the top-performing region from
Phase 1 and do a dense, high-accuracy sweep (n_samples=500) around
those best candidates only.

Early termination: Skip a candidate in Phase 2 if a first-pass quick
measurement (n_samples=50) comes back below a SNR threshold.

Output: results.csv (raw data) + results.xlsx (formatted Excel file)
  - Header row frozen so it stays visible while scrolling
  - Status column color coded (PASS=pastel green, FAIL=pastel red,
    OUT_OF_RANGE=pastel gray, SKIPPED=pastel yellow)
  - Top 3 best rows highlighted in bright neon green
  - Column widths auto-fitted
  - Summary row at the bottom showing the single best combination

Usage:
    python3 collect_data.py
"""

import csv
from device_sim import connect, configure, disconnect, measure

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
OUTPUT_CSV   = "results.csv"
OUTPUT_EXCEL = "results.xlsx"
FIELDNAMES   = ["trial", "phase", "status", "snr_db", "eye_height_mv",
                "eye_width_ps", "ber", "tx_level", "eq_gain", "pre_emphasis"]

# Phase 1: coarse grid
COARSE_TX      = range(100, 241, 35)
COARSE_EQ      = range(0, 16, 5)
COARSE_PRE     = range(0, 8, 3)
COARSE_SAMPLES = 100

# Phase 2: focused fine scan
FINE_SAMPLES       = 500
SNR_SKIP_THRESHOLD = 6.0
TOP_N_REGIONS      = 3
FINE_TX_RADIUS     = 25
FINE_EQ_RADIUS     = 2
FINE_PRE_RADIUS    = 1


def clamp(val, lo, hi):
    return max(lo, min(hi, val))


def build_fine_candidates(winners):
    seen = set()
    candidates = []
    for w in winners:
        tx0  = w["tx_level"]
        eq0  = w["eq_gain"]
        pre0 = w["pre_emphasis"]
        for dtx in range(-FINE_TX_RADIUS, FINE_TX_RADIUS + 1, 10):
            for deq in range(-FINE_EQ_RADIUS, FINE_EQ_RADIUS + 1):
                for dpre in range(-FINE_PRE_RADIUS, FINE_PRE_RADIUS + 1):
                    pt = (
                        clamp(tx0 + dtx,  0, 255),
                        clamp(eq0 + deq,  0,  15),
                        clamp(pre0 + dpre, 0,   7),
                    )
                    if pt not in seen:
                        seen.add(pt)
                        candidates.append({
                            "tx_level":     pt[0],
                            "eq_gain":      pt[1],
                            "pre_emphasis": pt[2],
                        })
    return candidates


def format_excel(all_rows):
    """
    Save results.xlsx with:
    - Frozen header row
    - Status color coding (pastel colors)
    - Top 3 best rows highlighted in bright neon green
    - Auto-fitted column widths
    - Bold summary row at the bottom
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # --- Header row ---
        ws.append(FIELDNAMES)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Freeze header row
        ws.freeze_panes = "A2"

        # --- Pastel fill colors for status ---
        fill_pass     = PatternFill("solid", fgColor="C6EFCE")  # pastel green
        fill_fail     = PatternFill("solid", fgColor="FFC7CE")  # pastel red
        fill_oor      = PatternFill("solid", fgColor="D9D9D9")  # pastel gray
        fill_skipped  = PatternFill("solid", fgColor="FFEB9C")  # pastel yellow

        # Neon green for top 3
        fill_neon     = PatternFill("solid", fgColor="00FF00")  # bright neon green

        # Find top 3 best rows by SNR
        passing = [r for r in all_rows if r.get("snr_db") not in (None, "")]
        passing_sorted = sorted(passing,
                                key=lambda r: float(r["snr_db"]),
                                reverse=True)
        top3_trials = set()
        for r in passing_sorted[:3]:
            top3_trials.add(r["trial"])

        # --- Data rows ---
        for row in all_rows:
            values = [row.get(f, "") for f in FIELDNAMES]
            ws.append(values)
            excel_row = ws.max_row
            trial_num = row.get("trial")
            status    = row.get("status", "")

            if trial_num in top3_trials:
                row_fill = fill_neon
            elif status == "PASS":
                row_fill = fill_pass
            elif status == "FAIL":
                row_fill = fill_fail
            elif status == "OUT_OF_RANGE":
                row_fill = fill_oor
            elif status == "SKIPPED":
                row_fill = fill_skipped
            else:
                row_fill = None

            if row_fill:
                for cell in ws[excel_row]:
                    cell.fill = row_fill

        # --- Summary row at the bottom ---
        if passing_sorted:
            best = passing_sorted[0]
            ws.append([])  # blank spacer row
            summary = [
                "BEST",
                int(best.get("phase", "")),
                best.get("status", ""),
                float(best.get("snr_db", 0)),
                float(best.get("eye_height_mv", 0)),
                float(best.get("eye_width_ps", 0)),
                best.get("ber", ""),
                int(best.get("tx_level", 0)),
                int(best.get("eq_gain", 0)),
                int(best.get("pre_emphasis", 0)),
            ]
            ws.append(summary)
            summary_row = ws.max_row
            bold_neon = PatternFill("solid", fgColor="00FF00")
            for cell in ws[summary_row]:
                cell.fill  = bold_neon
                cell.font  = Font(bold=True)

        # --- Auto-fit column widths ---
        for col_idx, col_name in enumerate(FIELDNAMES, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(col_name)
            for cell in ws[col_letter]:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 4

        wb.save(OUTPUT_EXCEL)
        print(f"Excel file saved to {OUTPUT_EXCEL}")
        print("  - Header row frozen")
        print("  - Status rows color coded")
        print("  - Top 3 best rows highlighted in neon green")
        print("  - Column widths auto-fitted")
        print("  - Best combination summary row added at bottom")

    except ImportError:
        print("openpyxl not installed — skipping Excel formatting.")
        print("Run: pip install openpyxl")


# ---------------------------------------------------------------------------
# Data collection
# ---------------------------------------------------------------------------
conn         = connect("192.168.10.5")
trial        = 0
phase1_results = []
all_rows     = []

with open(OUTPUT_CSV, "w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
    writer.writeheader()

    print("=== Phase 1: Coarse sweep — finding the good region ===")
    coarse_candidates = [
        {"tx_level": tx, "eq_gain": eq, "pre_emphasis": em}
        for tx in COARSE_TX
        for eq in COARSE_EQ
        for em in COARSE_PRE
    ]

    for candidate in coarse_candidates:
        configure(conn, candidate)
        data = measure(conn, n_samples=COARSE_SAMPLES)

        if data is None:
            status = "FAIL"
            row = {"trial": trial, "phase": 1, "status": status, **candidate}
        else:
            status = "PASS" if data["snr_db"] >= 8.0 else "OUT_OF_RANGE"
            row = {"trial": trial, "phase": 1, "status": status,
                   **data, **candidate}
            phase1_results.append(row)

        writer.writerow(row)
        f.flush()
        all_rows.append(row)

        if data:
            print(f"Trial {trial:>3d} [Phase 1] | {status:<12} "
                  f"| tx={candidate['tx_level']:>3d} "
                  f"eq={candidate['eq_gain']:>2d} "
                  f"pre={candidate['pre_emphasis']} "
                  f"| SNR={data['snr_db']:.2f} dB")
        else:
            print(f"Trial {trial:>3d} [Phase 1] | FAIL")
        trial += 1

    phase1_results.sort(key=lambda r: r.get("snr_db", -999), reverse=True)
    winners = phase1_results[:TOP_N_REGIONS]

    print(f"\nTop {TOP_N_REGIONS} results from Phase 1:")
    for w in winners:
        print(f"  tx={w['tx_level']} eq={w['eq_gain']} "
              f"pre={w['pre_emphasis']} "
              f"=> SNR={w.get('snr_db','?'):.2f} dB")

    print("\n=== Phase 2: Fine sweep — zooming into the best region ===")
    fine_candidates = build_fine_candidates(winners)

    for candidate in fine_candidates:
        configure(conn, candidate)

        quick = measure(conn, n_samples=50)
        if quick is not None and quick["snr_db"] < SNR_SKIP_THRESHOLD:
            status = "SKIPPED"
            row = {"trial": trial, "phase": 2, "status": status, **candidate}
            writer.writerow(row)
            f.flush()
            all_rows.append(row)
            print(f"Trial {trial:>3d} [Phase 2] | SKIPPED "
                  f"(SNR={quick['snr_db']:.1f} dB — too low)")
            trial += 1
            continue

        data = measure(conn, n_samples=FINE_SAMPLES)
        if data is None:
            status = "FAIL"
            row = {"trial": trial, "phase": 2, "status": status, **candidate}
        else:
            status = "PASS" if data["snr_db"] >= 8.0 else "OUT_OF_RANGE"
            row = {"trial": trial, "phase": 2, "status": status,
                   **data, **candidate}

        writer.writerow(row)
        f.flush()
        all_rows.append(row)

        if data:
            print(f"Trial {trial:>3d} [Phase 2] | {status:<12} "
                  f"| tx={candidate['tx_level']:>3d} "
                  f"eq={candidate['eq_gain']:>2d} "
                  f"pre={candidate['pre_emphasis']} "
                  f"| SNR={data['snr_db']:.2f} dB")
        else:
            print(f"Trial {trial:>3d} [Phase 2] | FAIL")
        trial += 1

disconnect(conn)
print(f"\nDone. {trial} total trials. Results saved to {OUTPUT_CSV}")

# Print best result
best = max(phase1_results, key=lambda r: r.get("snr_db", -999))
print(f"\nBest setting found:")
print(f"  tx_level={best['tx_level']}  "
      f"eq_gain={best['eq_gain']}  "
      f"pre_emphasis={best['pre_emphasis']}")
print(f"  SNR = {best['snr_db']:.2f} dB")

# Generate formatted Excel file
print(f"\nGenerating formatted Excel file...")
format_excel(all_rows)
